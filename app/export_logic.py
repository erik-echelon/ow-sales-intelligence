"""
Business Logic for Excel Export

Contains functions for:
- Merging data from multiple sources (scored, research, raw companies)
- Formatting data for export
- Generating Excel files with multiple sheets
- JSON field expansion (hot_lead_signals, concerns, research_summary)
"""

import io
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def merge_export_data(
    scored_df: pd.DataFrame,
    research_df: Optional[pd.DataFrame],
    companies_df: pd.DataFrame,
    company_ids: List[str],
    hubspot_flags: Dict[str, bool]
) -> pd.DataFrame:
    """
    Merge all data sources for selected companies.

    Args:
        scored_df: Scored companies dataframe
        research_df: Company research dataframe (may be None)
        companies_df: Raw companies dataframe
        company_ids: List of company IDs to export
        hubspot_flags: Dict mapping company_id to in_hubspot boolean

    Returns:
        Merged dataframe with all data for export
    """
    # Filter to selected companies
    merged_df = scored_df[scored_df['company_id'].astype(str).isin(company_ids)].copy()

    # Add HubSpot flags
    merged_df['in_hubspot'] = merged_df['company_id'].astype(str).map(hubspot_flags).fillna(False)

    # Merge research data if available
    if research_df is not None and len(research_df) > 0:
        research_cols = ['company_id', 'had_web_research', 'confidence', 'reasoning',
                        'hot_lead_signals', 'concerns', 'recommendation', 'timing', 'research_summary']
        # Only include columns that exist
        research_cols = [col for col in research_cols if col in research_df.columns]

        merged_df = merged_df.merge(
            research_df[research_cols],
            on='company_id',
            how='left',
            suffixes=('', '_research')
        )

    # Merge raw company data for additional fields
    raw_cols = ['company_id', 'website', 'building_count_estimate', 'growing_business_code', 'contacts_count']
    # Only include columns that exist in companies_df
    raw_cols = [col for col in raw_cols if col in companies_df.columns]

    # Remove duplicates from companies_df before merge
    companies_unique = companies_df[raw_cols].drop_duplicates(subset=['company_id'], keep='first')

    merged_df = merged_df.merge(
        companies_unique,
        on='company_id',
        how='left',
        suffixes=('', '_raw')
    )

    # Remove any duplicate rows that may have been created during merging
    merged_df = merged_df.drop_duplicates(subset=['company_id'], keep='first')

    return merged_df


def expand_json_array(json_str: Optional[str]) -> tuple[int, str]:
    """
    Expand JSON array to count and formatted string.

    Args:
        json_str: JSON array string like '["item1", "item2"]'

    Returns:
        Tuple of (count, formatted_string)
    """
    if pd.isna(json_str) or not json_str:
        return 0, ""

    try:
        items = json.loads(json_str)
        if isinstance(items, list):
            count = len(items)
            formatted = "\n".join(items)
            return count, formatted
        else:
            return 0, str(items)
    except (json.JSONDecodeError, TypeError):
        return 0, str(json_str)


def expand_research_summary(json_str: Optional[str]) -> Dict[str, Any]:
    """
    Parse and flatten research_summary JSON.

    Args:
        json_str: JSON string containing research summary

    Returns:
        Flattened dictionary with research details
    """
    result = {
        'news_found': False,
        'news_headline': '',
        'news_date': '',
        'news_signal_type': '',
        'news_summary': '',
        'jobs_found_ops_roles': False,
        'jobs_hiring_scale': '',
        'jobs_role_examples': '',
        'corp_has_vendor_portal': False,
        'corp_has_facilities_dept': False,
        'corp_centralized_operations': False,
        'corp_professionalism_level': '',
        'corp_summary': ''
    }

    if pd.isna(json_str) or not json_str:
        return result

    try:
        summary = json.loads(json_str)

        # News findings
        if 'news_findings' in summary:
            news = summary['news_findings']
            result['news_found'] = news.get('found_recent_news', False)
            result['news_headline'] = news.get('headline', '')
            result['news_date'] = news.get('date', '')
            result['news_signal_type'] = news.get('signal_type', '')
            result['news_summary'] = news.get('summary', '')

        # Jobs findings
        if 'jobs_findings' in summary:
            jobs = summary['jobs_findings']
            result['jobs_found_ops_roles'] = jobs.get('found_ops_roles', False)
            result['jobs_hiring_scale'] = jobs.get('hiring_scale', '')
            role_examples = jobs.get('role_examples', [])
            if isinstance(role_examples, list):
                result['jobs_role_examples'] = "\n".join(role_examples[:5])  # First 5 examples
            else:
                result['jobs_role_examples'] = str(role_examples)

        # Corporate ops findings
        if 'corporate_ops_findings' in summary:
            corp = summary['corporate_ops_findings']
            result['corp_has_vendor_portal'] = corp.get('has_vendor_portal', False)
            result['corp_has_facilities_dept'] = corp.get('has_facilities_dept', False)
            result['corp_centralized_operations'] = corp.get('centralized_operations', False)
            result['corp_professionalism_level'] = corp.get('professionalism_level', '')
            result['corp_summary'] = corp.get('summary', '')

    except (json.JSONDecodeError, TypeError) as e:
        logger.warning(f"Failed to parse research_summary: {e}")

    return result


def prepare_sheet1_company_rankings(merged_df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare data for Sheet 1: Company Rankings.

    Args:
        merged_df: Merged dataframe with all data

    Returns:
        Formatted dataframe for Sheet 1
    """
    sheet1_df = merged_df.copy()

    # Expand JSON fields for summary counts
    if 'hot_lead_signals' in sheet1_df.columns:
        sheet1_df[['signals_count', 'signals_text']] = sheet1_df['hot_lead_signals'].apply(
            lambda x: pd.Series(expand_json_array(x))
        )
    else:
        sheet1_df['signals_count'] = 0
        sheet1_df['signals_text'] = ''

    if 'concerns' in sheet1_df.columns:
        sheet1_df[['concerns_count', 'concerns_text']] = sheet1_df['concerns'].apply(
            lambda x: pd.Series(expand_json_array(x))
        )
    else:
        sheet1_df['concerns_count'] = 0
        sheet1_df['concerns_text'] = ''

    # Truncate ICP reasoning for main sheet
    if 'reasoning' in sheet1_df.columns:
        sheet1_df['icp_reasoning_short'] = sheet1_df['reasoning'].apply(
            lambda x: (str(x)[:500] + '...') if pd.notna(x) and len(str(x)) > 500 else str(x) if pd.notna(x) else ''
        )
    elif 'icp_fit_reasoning' in sheet1_df.columns:
        sheet1_df['icp_reasoning_short'] = sheet1_df['icp_fit_reasoning'].apply(
            lambda x: (str(x)[:500] + '...') if pd.notna(x) and len(str(x)) > 500 else str(x) if pd.notna(x) else ''
        )
    else:
        sheet1_df['icp_reasoning_short'] = ''

    # Select and order columns for Sheet 1
    columns_order = [
        # Identification
        'company_id', 'company_name', 'primary_naics', 'naics_4digit', 'city', 'state',
        # Status
        'scoring_path', 'is_customer', 'in_hubspot',
        # Scores
        'final_score', 'naics_attractiveness_score', 'company_opportunity_score',
        # Prospect component scores
        'icp_fit_score', 'buildings_score', 'revenue_score', 'growth_score', 'contact_score',
        # Customer component scores
        'expansion_score', 'churn_score', 'profitability_score', 'tickets_score',
        # Raw data
        'location_employee_size', 'sales_volume', 'building_count_estimate',
        'growing_business_code', 'contacts_count', 'website',
        # ICP assessment
        'recommendation', 'confidence', 'icp_reasoning_short',
        # Signals
        'signals_count', 'concerns_count', 'timing',
        # Metadata
        'rank', 'scoring_reason', 'has_research_doc', 'research_recency'
    ]

    # Only include columns that exist
    columns_order = [col for col in columns_order if col in sheet1_df.columns]

    result_df = sheet1_df[columns_order].copy()

    # Rename columns for better readability
    column_renames = {
        'company_id': 'Company ID',
        'company_name': 'Company Name',
        'primary_naics': 'NAICS Code (8-digit)',
        'naics_4digit': 'NAICS (4-digit)',
        'city': 'City',
        'state': 'State',
        'scoring_path': 'Scoring Path',
        'is_customer': 'Is Customer',
        'in_hubspot': 'In HubSpot',
        'final_score': 'Final Score',
        'naics_attractiveness_score': 'NAICS Score',
        'company_opportunity_score': 'Company Score',
        'icp_fit_score': 'ICP Fit Score',
        'buildings_score': 'Buildings Score',
        'revenue_score': 'Revenue Score',
        'growth_score': 'Employee Score',
        'contact_score': 'Contact Score',
        'expansion_score': 'Expansion Score',
        'churn_score': 'Churn Score',
        'profitability_score': 'Profitability Score',
        'tickets_score': 'Tickets Score',
        'location_employee_size': 'Employees',
        'sales_volume': 'Revenue ($)',
        'building_count_estimate': 'Building Count Estimate',
        'growing_business_code': 'Growth Code',
        'contacts_count': 'Contact Count',
        'website': 'Website',
        'recommendation': 'ICP Recommendation',
        'confidence': 'ICP Confidence',
        'icp_reasoning_short': 'ICP Reasoning (Truncated)',
        'signals_count': 'Hot Lead Signals (Count)',
        'concerns_count': 'Concerns (Count)',
        'timing': 'Timing Assessment',
        'rank': 'Segment Rank',
        'scoring_reason': 'Scoring Reason',
        'has_research_doc': 'Has Research',
        'research_recency': 'Research Recency'
    }

    result_df = result_df.rename(columns=column_renames)

    return result_df


def prepare_sheet2_icp_reasoning(merged_df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare data for Sheet 2: ICP Detailed Reasoning.

    Args:
        merged_df: Merged dataframe with all data

    Returns:
        Formatted dataframe for Sheet 2
    """
    sheet2_df = merged_df.copy()

    # Expand JSON arrays
    if 'hot_lead_signals' in sheet2_df.columns:
        sheet2_df[['signals_count', 'hot_lead_signals_expanded']] = sheet2_df['hot_lead_signals'].apply(
            lambda x: pd.Series(expand_json_array(x))
        )
    else:
        sheet2_df['hot_lead_signals_expanded'] = ''

    if 'concerns' in sheet2_df.columns:
        sheet2_df[['concerns_count', 'concerns_expanded']] = sheet2_df['concerns'].apply(
            lambda x: pd.Series(expand_json_array(x))
        )
    else:
        sheet2_df['concerns_expanded'] = ''

    # Select columns
    columns = [
        'company_id', 'company_name', 'icp_fit_score', 'recommendation', 'confidence',
        'reasoning', 'hot_lead_signals_expanded', 'concerns_expanded', 'timing'
    ]

    # Use alternative column names if primary not available
    if 'reasoning' not in sheet2_df.columns and 'icp_fit_reasoning' in sheet2_df.columns:
        sheet2_df['reasoning'] = sheet2_df['icp_fit_reasoning']

    columns = [col for col in columns if col in sheet2_df.columns]
    result_df = sheet2_df[columns].copy()

    # Rename columns
    column_renames = {
        'company_id': 'Company ID',
        'company_name': 'Company Name',
        'icp_fit_score': 'ICP Fit Score',
        'recommendation': 'Recommendation',
        'confidence': 'Confidence',
        'reasoning': 'ICP Fit Reasoning (Full)',
        'hot_lead_signals_expanded': 'Hot Lead Signals',
        'concerns_expanded': 'Concerns',
        'timing': 'Timing Assessment'
    }

    result_df = result_df.rename(columns=column_renames)

    return result_df


def prepare_sheet3_research_details(merged_df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare data for Sheet 3: Web Research Details.

    Args:
        merged_df: Merged dataframe with all data

    Returns:
        Formatted dataframe for Sheet 3
    """
    # Only include companies that have research
    if 'had_web_research' in merged_df.columns:
        sheet3_df = merged_df[merged_df['had_web_research'] == True].copy()
    else:
        # Return empty dataframe if no research column
        return pd.DataFrame()

    if len(sheet3_df) == 0:
        return pd.DataFrame()

    # Expand research_summary JSON
    if 'research_summary' in sheet3_df.columns:
        research_expanded = sheet3_df['research_summary'].apply(expand_research_summary)
        research_df = pd.DataFrame(research_expanded.tolist())

        # Combine with base data
        result_df = pd.concat([
            sheet3_df[['company_id', 'company_name', 'had_web_research']].reset_index(drop=True),
            research_df.reset_index(drop=True)
        ], axis=1)
    else:
        result_df = sheet3_df[['company_id', 'company_name', 'had_web_research']].copy()

    # Rename columns
    column_renames = {
        'company_id': 'Company ID',
        'company_name': 'Company Name',
        'had_web_research': 'Had Web Research',
        'news_found': 'News Found',
        'news_headline': 'News Headline',
        'news_date': 'News Date',
        'news_signal_type': 'News Signal Type',
        'news_summary': 'News Summary',
        'jobs_found_ops_roles': 'Found Ops Roles',
        'jobs_hiring_scale': 'Hiring Scale',
        'jobs_role_examples': 'Job Role Examples',
        'corp_has_vendor_portal': 'Has Vendor Portal',
        'corp_has_facilities_dept': 'Has Facilities Dept',
        'corp_centralized_operations': 'Centralized Operations',
        'corp_professionalism_level': 'Professionalism Level',
        'corp_summary': 'Corporate Ops Summary'
    }

    result_df = result_df.rename(columns=column_renames)

    return result_df


def create_export_excel(
    merged_df: pd.DataFrame,
    filter_info: Dict[str, Any],
    output: Union[Path, io.BytesIO]
) -> None:
    """
    Generate multi-sheet Excel file with formatting.

    Args:
        merged_df: Merged dataframe with all data
        filter_info: Dictionary containing filter criteria applied
        output: Output path or BytesIO buffer
    """
    # Prepare data for each sheet
    sheet1_df = prepare_sheet1_company_rankings(merged_df)
    sheet2_df = prepare_sheet2_icp_reasoning(merged_df)
    sheet3_df = prepare_sheet3_research_details(merged_df)

    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write Sheet 1: Company Rankings
        sheet1_df.to_excel(writer, sheet_name='Company Rankings', index=False)

        # Write Sheet 2: ICP Detailed Reasoning
        if len(sheet2_df) > 0:
            sheet2_df.to_excel(writer, sheet_name='ICP Detailed Reasoning', index=False)

        # Write Sheet 3: Web Research Details (only if there's data)
        if len(sheet3_df) > 0:
            sheet3_df.to_excel(writer, sheet_name='Web Research Details', index=False)

        # Write Sheet 4: Export Metadata
        metadata_df = pd.DataFrame({
            'Property': [
                'Export Timestamp',
                'Total Companies Exported',
                'Filter: Customer Status',
                'Filter: NAICS Code',
                'Filter: HubSpot Status',
                'Filter: State',
                'Filter: Score Range (Min)',
                'Filter: Score Range (Max)'
            ],
            'Value': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                len(merged_df),
                filter_info.get('customer_status', 'All'),
                filter_info.get('naics_code', 'All'),
                filter_info.get('hubspot_status', 'All'),
                filter_info.get('state', 'All'),
                filter_info.get('score_min', 0),
                filter_info.get('score_max', 100)
            ]
        })
        metadata_df.to_excel(writer, sheet_name='Export Metadata', index=False)

        # Apply formatting
        workbook = writer.book
        format_worksheet(workbook['Company Rankings'], sheet1_df)
        if 'ICP Detailed Reasoning' in workbook.sheetnames:
            format_worksheet(workbook['ICP Detailed Reasoning'], sheet2_df)
        if 'Web Research Details' in workbook.sheetnames:
            format_worksheet(workbook['Web Research Details'], sheet3_df)
        format_worksheet(workbook['Export Metadata'], metadata_df)


def format_worksheet(worksheet: Worksheet, df: pd.DataFrame) -> None:
    """
    Apply formatting to worksheet.

    Args:
        worksheet: openpyxl Worksheet object
        df: DataFrame (for conditional formatting on Final Score)
    """
    # Header formatting
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Freeze first row
    worksheet.freeze_panes = 'A2'

    # Auto-size columns (with max width)
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Set column width (max 50 chars)
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Apply conditional formatting to Final Score column if it exists
    if 'Final Score' in df.columns:
        final_score_col_idx = df.columns.get_loc('Final Score') + 1  # Excel is 1-indexed
        final_score_col_letter = worksheet.cell(row=1, column=final_score_col_idx).column_letter

        # Define color fills
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        light_green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        orange_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Apply conditional formatting to data rows
        for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
            cell = worksheet[f'{final_score_col_letter}{row_idx}']

            try:
                score = float(cell.value) if cell.value else 0

                if score >= 90:
                    cell.fill = green_fill
                elif score >= 80:
                    cell.fill = light_green_fill
                elif score >= 70:
                    cell.fill = yellow_fill
                elif score >= 60:
                    cell.fill = orange_fill
                else:
                    cell.fill = red_fill
            except (ValueError, TypeError):
                pass
